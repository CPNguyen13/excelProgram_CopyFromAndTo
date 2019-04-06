import os
import xlrd
import openpyxl
import glob
import time

def copytotarget(copyfrom, copyto, days_ago):
    """
        :param copyfrom: file to copy from
        :param copyto: target file
        :param days_ago: how many days was the previous workday from today.
        :return: NONE, it copies the contents of copyfrom to copyto based on previous workday.

        Example: if today was 1/12/2019 and days_ago = 10, all records from everyone will be copied to copyto file
        with the date 1/2/2019.
    """

    def datetofind(lastworkday):
        """
            This function makes sure time in which to copy is correct, takes into account epoch time, timezone, etc.
            :param lastwordday: same as days_ago (how many days was the previous workday from today.)
            :return: Returns back the number of days since epoch and timezone adjustments.
        """
        sec_from_1900 = 2209161600 + time.mktime(time.localtime()) - ((lastworkday) * 86400) - 21600
        return sec_from_1900 // 86400

    global col_idx_pairs
    wbread = xlrd.open_workbook(copyfrom)
    wsread = wbread.sheet_by_index(0)
    daterowindex = 2

    # This is to locate the last row and append after that.
    last_row = xlrd.open_workbook(copyto).sheet_by_index(0).nrows
    copyfromlastrow = wsread.nrows

    while daterowindex < copyfromlastrow:
        if wsread.cell_value(daterowindex, 0) == datetofind(days_ago):
            for item in col_idx_pairs:  # This will copy an entire row/record.
                if wsread.cell_value(daterowindex, item[1]) == 42:
                    ws['{0}{1}'.format(item[0], last_row + 1)] = '#N/A'
                else:
                    ws['{0}{1}'.format(item[0], last_row + 1)] = wsread.cell_value(daterowindex, item[1])
            last_row += 1
        daterowindex += 1
    wb.save(copyto)



# Change this section based on used Excel columns/fields.
# a(0) b(1) c d(3) e(4) f(5) g h i j k(10) l m(12) n(13) o p(15) q(16) r(17) s(18) t(19) u(20) v
columns = ['A', 'B', 'D', 'E', 'F', 'K', 'M', 'N', 'P', 'Q', 'R', 'S', 'T', 'U']
indexes = [0, 1, 3, 4, 5, 10, 12, 13, 15, 16, 17, 18, 19, 20]
col_idx_pairs = list(zip(columns, indexes))

# Set up directory and files
directory = r''  # Change as needed, this is hardcoded for now.
os.chdir(directory)
target = 'OOW_MAIN_2018.xlsx'  # Change as needed
wb = openpyxl.load_workbook(target)
ws = wb.active

# Get everyone's file
readfiles = glob.glob('OOW_*_2018.xlsx')
readfiles.remove('OOW_MAIN_2018.xlsx')

#Makes sure days_ago is a valid value (integer), not complete
days_ago = int(input("How many days ago was the last workday?: "))

#The beginning of the copying process
print("Starting Process.")
for file in readfiles:
    copytotarget(file, target, days_ago)
print('Process Completed!')
